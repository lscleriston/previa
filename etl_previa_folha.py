import sqlite3
import os
import openpyxl

DB_PATH = os.environ.get("DB_PATH", os.path.join('data', 'previadb.db'))
XLSX_PATH = os.environ.get("XLSX_PATH", os.path.join('data', 'Forecast Semanal 2026 - Abril.xlsx'))


def create_table_if_not_exists(conn):
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS previa_folha_th (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mes_ref TEXT,
            cr TEXT,
            cliente TEXT,
            valor REAL,
            fonte TEXT
        )
    ''')
    conn.commit()


def parse_float(val):
    if val is None:
        return 0.0
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).replace('R$', '').replace(' ', '').replace(',', '.')
        return float(s)
    except Exception:
        return 0.0


def run_etl():
    print(f"Iniciando ETL de Prévia Folha TH Abr a partir de {XLSX_PATH}")
    if not os.path.exists(XLSX_PATH):
        print(f"Arquivo não encontrado: {XLSX_PATH}")
        return

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    sheet_name = None
    for s in wb.sheetnames:
        if s == 'Prévia Folha TH Abr':
            sheet_name = s
            break
    if not sheet_name:
        print('Aba Prévia Folha TH Abr não encontrada.')
        return

    sheet = wb[sheet_name]
    mes_ref = '2026-04'

    conn = sqlite3.connect(DB_PATH)
    create_table_if_not_exists(conn)
    cursor = conn.cursor()

    cursor.execute('DELETE FROM previa_folha_th WHERE mes_ref = ?', (mes_ref,))
    conn.commit()

    count = 0
    for row in sheet.iter_rows(min_row=3, values_only=True):
        cliente = str(row[0]).strip() if row[0] is not None else None
        cr = str(row[1]).strip() if row[1] is not None else None
        if not cr or cr.lower() == 'cr':
            continue

        valor = 0.0
        for cell in row[2:]:
            valor += parse_float(cell)

        if valor == 0.0:
            continue

        cursor.execute('''
            INSERT INTO previa_folha_th (mes_ref, cr, cliente, valor, fonte)
            VALUES (?, ?, ?, ?, ?)
        ''', (
            mes_ref,
            cr,
            cliente,
            valor,
            'Prévia Folha TH Abr'
        ))
        count += 1

    conn.commit()
    conn.close()
    print(f"ETL Prévia Folha TH Abr concluído. {count} registros inseridos.")


if __name__ == '__main__':
    run_etl()
