import sqlite3
import os
import openpyxl
from datetime import datetime

DB_PATH = os.environ.get(
    "DB_PATH",
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "db", "previadb.db")
)
XLSX_PATH = os.environ.get(
    "XLSX_PATH",
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "raw", "Forecast Semanal 2026 - Abril.xlsx")
)

def create_table_if_not_exists(conn):
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ajustamentos_gerencia (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            gerencia TEXT,
            resultado TEXT,
            cr_credito TEXT,
            cr_debito TEXT,
            mes_ref TEXT,
            incremento_credito REAL,
            justificativa TEXT,
            incremento_debito REAL,
            cr_envio TEXT,
            desc_cr_envio TEXT,
            gestor_cr_envio TEXT,
            cr_destino TEXT,
            desc_cr_destino TEXT,
            gestor_cr_destino TEXT
        )
    ''')
    conn.commit()

def parse_float(val):
    if val is None:
        return 0.0
    try:
        if isinstance(val, (int, float)):
            return float(val)
        # remove possible R$, spaces, etc
        s = str(val).replace('R$', '').replace(' ', '').replace(',', '.')
        return float(s)
    except:
        return 0.0

def load_sheet(conn, wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        print(f"Aba {sheet_name} não encontrada.")
        return
        
    print(f"Carregando aba: {sheet_name}")
    sheet = wb[sheet_name]
    cursor = conn.cursor()
    
    count = 0
    
    # "GERENCIA OCTAVIO" and "GERENCIA WESLEY" have different columns
    if sheet_name == "GERENCIA EDILSON":
        # starts at row 5
        for row in sheet.iter_rows(min_row=5, values_only=True):
            if not row[0]: continue
            resultado = str(row[0]) if row[0] is not None else None
            cr_credito = str(row[1]) if row[1] is not None else None
            cr_debito = str(row[2]) if row[2] is not None else None
            
            mes_ref = None
            if isinstance(row[3], datetime):
                mes_ref = row[3].strftime('%Y-%m')
            elif row[3]:
                try: mes_ref = str(row[3])[:7]
                except: mes_ref = str(row[3])
                    
            incremento_credito = parse_float(row[4])
            justificativa = str(row[5]) if row[5] is not None else None
            incremento_debito = parse_float(row[6])
            cr_envio = str(row[7]) if row[7] is not None else None
            desc_cr_envio = str(row[8]) if row[8] is not None else None
            gestor_cr_envio = str(row[9]) if row[9] is not None else None
            cr_destino = str(row[10]) if row[10] is not None else None
            desc_cr_destino = str(row[11]) if row[11] is not None else None
            gestor_cr_destino = str(row[12]) if row[12] is not None else None
            
            cursor.execute('''
                INSERT INTO ajustamentos_gerencia (
                    gerencia, resultado, cr_credito, cr_debito, mes_ref,
                    incremento_credito, justificativa, incremento_debito,
                    cr_envio, desc_cr_envio, gestor_cr_envio,
                    cr_destino, desc_cr_destino, gestor_cr_destino
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                sheet_name, resultado, cr_credito, cr_debito, mes_ref,
                incremento_credito, justificativa, incremento_debito,
                cr_envio, desc_cr_envio, gestor_cr_envio,
                cr_destino, desc_cr_destino, gestor_cr_destino
            ))
            count += 1
            
    elif sheet_name == "GERENCIA OCTAVIO":
        # ('Grupo de Conta', 'CR CRÉDITO', 'CR DÉBITO', 'Mês', 'INCREMENTO CRÉDITO', 'INCREMENTO DÉBITO', 'Coordenador', 'Observação')
        # starts at row 3
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row[0]: continue
            resultado = str(row[0]) if row[0] is not None else None # 'Grupo de Conta' -> resultado
            cr_credito = str(row[1]) if row[1] is not None else None
            cr_debito = str(row[2]) if row[2] is not None else None
            
            mes_ref = None
            if isinstance(row[3], datetime):
                mes_ref = row[3].strftime('%Y-%m')
            elif row[3]:
                try: mes_ref = str(row[3])[:7]
                except: mes_ref = str(row[3])
                
            incremento_credito = parse_float(row[4])
            incremento_debito = parse_float(row[5])
            
            # Map Coordenador to desc / justificativa to observação, ou whatever makes sense, keeping it simple here
            # I will map 'Coordenador' to gestor_cr_destino for reference, and 'Observacao' to justificativa
            gestor_cr_destino = str(row[6]) if row[6] is not None else None
            justificativa = str(row[7]) if row[7] is not None else None
            
            cursor.execute('''
                INSERT INTO ajustamentos_gerencia (
                    gerencia, resultado, cr_credito, cr_debito, mes_ref,
                    incremento_credito, justificativa, incremento_debito,
                    cr_envio, desc_cr_envio, gestor_cr_envio,
                    cr_destino, desc_cr_destino, gestor_cr_destino
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                sheet_name, resultado, cr_credito, cr_debito, mes_ref,
                incremento_credito, justificativa, incremento_debito,
                None, None, None,
                None, None, gestor_cr_destino
            ))
            count += 1
            
    elif sheet_name == "GERENCIA WESLEY":
        # Similar to OCTAVIO based on the user request usually. Let's make it robust by trying to guess indices
        # If it's same as Edilson it falls back to parsing cleanly if we are careful. Let's do the safe way:
        for i, row in enumerate(sheet.iter_rows(min_row=1, values_only=True)):
            if i < 2: continue # skip header heuristic
            if len(row) < 5: continue
            if not row[0]: continue
            if str(row[0]).upper() in ["RESULTADO", "GRUPO DE CONTA"]: continue # skip header line
            
            # let's assume it's like OCTAVIO if len is around 8-10, or like EDILSON if len >= 13
            is_edilson_like = len(row) >= 13
            
            resultado = str(row[0]) if row[0] is not None else None
            cr_credito = str(row[1]) if row[1] is not None else None
            cr_debito = str(row[2]) if row[2] is not None else None
            
            mes_ref = None
            if isinstance(row[3], datetime):
                mes_ref = row[3].strftime('%Y-%m')
            elif row[3]:
                try: mes_ref = str(row[3])[:7]
                except: mes_ref = str(row[3])
                
            incremento_credito = parse_float(row[4])
            
            if is_edilson_like:
                justificativa = str(row[5]) if row[5] is not None else None
                incremento_debito = parse_float(row[6])
                gestor_cr_destino = str(row[12]) if row[12] is not None else None
            else:
                try:
                    incremento_debito = parse_float(row[5])
                    gestor_cr_destino = str(row[6]) if len(row) > 6 and row[6] else None
                    justificativa = str(row[7]) if len(row) > 7 and row[7] else None
                except:
                    incremento_debito = 0.0
                    gestor_cr_destino = None
                    justificativa = None
            
            cursor.execute('''
                INSERT INTO ajustamentos_gerencia (
                    gerencia, resultado, cr_credito, cr_debito, mes_ref,
                    incremento_credito, justificativa, incremento_debito,
                    cr_envio, desc_cr_envio, gestor_cr_envio,
                    cr_destino, desc_cr_destino, gestor_cr_destino
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                sheet_name, resultado, cr_credito, cr_debito, mes_ref,
                incremento_credito, justificativa, incremento_debito,
                None, None, None,
                None, None, gestor_cr_destino
            ))
            count += 1
            
    conn.commit()
    print(f"Abas '{sheet_name}': Inseridos {count} registros.")

if __name__ == "__main__":
    print(f"Iniciando ETL das gerencias a partir de {XLSX_PATH}")
    
    if not os.path.exists(XLSX_PATH):
        print(f"Arquivo não encontrado: {XLSX_PATH}")
        exit(1)
        
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    
    conn = sqlite3.connect(DB_PATH)
    create_table_if_not_exists(conn)
    
    # Limpar a tabela antes de popular
    conn.execute("DELETE FROM ajustamentos_gerencia")
    conn.commit()
    
    gerencias = ["GERENCIA EDILSON", "GERENCIA OCTAVIO", "GERENCIA WESLEY"]
    
    for g in gerencias:
        load_sheet(conn, wb, g)
        
    conn.close()
    print("ETL concluído com sucesso.")
