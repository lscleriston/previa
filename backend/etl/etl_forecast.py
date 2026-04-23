import sqlite3
import openpyxl
import datetime
import os
import glob
import re

DB_PATH = os.environ.get(
    "DB_PATH",
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "db", "previadb.db")
)
XLSX_PATH = os.environ.get(
    "XLSX_PATH",
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "raw", "Forecast Semanal 2026 - Abril.xlsx")
)

def clear_db():
    pass

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.executescript("""
    CREATE TABLE IF NOT EXISTS forecast_oportunidades (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chave_ek TEXT UNIQUE,
        data_criacao DATE,
        banco TEXT,
        ano INTEGER,
        pais TEXT,
        tipo_papel TEXT,
        gerente TEXT,
        owner TEXT,
        ger_comercial TEXT,
        operacao TEXT,
        ger_operacao TEXT,
        pre_vendas TEXT,
        pratica TEXT,
        produto TEXT,
        id_empresa TEXT,
        cliente TEXT,
        cliente_ge TEXT,
        subcliente TEXT,
        novo_cliente BOOLEAN,
        industria TEXT,
        id_oportunidade TEXT,
        descricao_oportunidade TEXT,
        status_comercial TEXT,
        status_comercial_det TEXT,
        metodologia TEXT,
        tipo_opp TEXT,
        cr TEXT,
        cr2 TEXT,
        semana_fechamento TEXT,
        cr_oi TEXT,
        ordem_interna TEXT,
        moeda TEXT,
        consideracao TEXT,
        vigencia TEXT,
        contrato TEXT,
        data_inicio_contrato DATE,
        data_fim_contrato DATE,
        mes_reajuste TEXT,
        risco TEXT,
        deducao REAL,
        pct_ponderacao REAL,
        semana_carga DATETIME,
        arquivo_origem TEXT
    );

    CREATE TABLE IF NOT EXISTS forecast_valores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chave_ek TEXT,
        cenario TEXT,
        mes_ref TEXT,
        valor_rl REAL,
        valor_rb REAL,
        semana_carga DATETIME,
        FOREIGN KEY (chave_ek) REFERENCES forecast_oportunidades(chave_ek)
    );

    CREATE TABLE IF NOT EXISTS etl_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        arquivo TEXT,
        aba TEXT,
        linhas_lidas INTEGER,
        linhas_carregadas INTEGER,
        linhas_ignoradas INTEGER,
        status TEXT,
        mensagem TEXT,
        executado_em DATETIME
    );
    """)
    conn.commit()
    return conn


def ensure_forecast_oportunidades_schema(cursor):
    cursor.execute("PRAGMA table_info(forecast_oportunidades)")
    cols = [row[1] for row in cursor.fetchall()]
    if 'cr2' not in cols:
        cursor.execute("ALTER TABLE forecast_oportunidades ADD COLUMN cr2 TEXT")


def parse_date(val):
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    return None

def parse_float(val):
    if not val:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0


def parse_cr_values(cr_raw):
    if cr_raw is None:
        return None, None
    text = str(cr_raw).strip()
    if not text or text.lower() == 'none':
        return None, None
    parts = [part.strip() for part in re.split(r'[-/,;]', text) if part and part.strip().lower() != 'none']
    cr = parts[0] if len(parts) > 0 else None
    cr2 = parts[1] if len(parts) > 1 else None
    return cr, cr2


def normalize_header_name(val):
    if val is None:
        return ""
    text = str(val).strip().lower()
    for src, dst in [('_', ' '), ('-', ' '), ('ç', 'c'), ('é', 'e'), ('á', 'a'), ('ã', 'a'), ('õ', 'o'), ('í', 'i'), ('ó', 'o'), ('ú', 'u')]:
        text = text.replace(src, dst)
    return text


def find_header_index(headers, candidates):
    for idx, header in enumerate(headers):
        if not header:
            continue
        for candidate in candidates:
            if candidate in header:
                return idx
    return None


def find_forecast_header(sheet, max_rows=20):
    for row_idx in range(1, max_rows + 1):
        row = list(sheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
        normalized = [normalize_header_name(cell) for cell in row]
        if (find_header_index(normalized, ['chave ek', 'chave_ek', 'ek', 'e k', 'chave']) is not None and
                find_header_index(normalized, ['cr sap', 'cr_sap', 'cr', 'cod cr', 'cod_cr']) is not None):
            return row_idx, normalized
    fallback = list(sheet.iter_rows(min_row=4, max_row=4, values_only=True))[0]
    return 4, [normalize_header_name(cell) for cell in fallback]


def get_cell(row, idx, default=None):
    if idx is None or idx < 0:
        return default
    return row[idx] if len(row) > idx else default


def run_dim_cr_etl(wb, cursor, agora):
    print("Iniciando carga da Dimensão CR...")
    sheet_name = None
    for s in wb.sheetnames:
        if s.startswith('Prévia') and 'MSP' in s:
            sheet_name = s
            break
            
    if not sheet_name:
        print("Aba de Prévia não encontrada para carga da dim_cr.")
        return

    print(f"Lendo Dim CR da aba: {sheet_name}")
    sheet = wb[sheet_name]
    cursor.execute("PRAGMA table_info(dim_cr)")
    existing_cols = [row[1] for row in cursor.fetchall()]
    if 'CR_SAP' not in existing_cols:
        cursor.execute("ALTER TABLE dim_cr ADD COLUMN CR_SAP TEXT")
        cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_dim_cr_cr_sap ON dim_cr(CR_SAP)")
    
    header_row = list(sheet.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    
    try:
        idx_cr_sap = header_row.index('CR SAP')
        idx_cod_cr = header_row.index('Cod_Cr')
        idx_cliente = header_row.index('Cliente')
        idx_des_cr = header_row.index('Des_CR')
        idx_pais = header_row.index('Pais')
        idx_diretor = header_row.index('Diretor')
        idx_gerente = header_row.index('Gerente')
    except ValueError as e:
        print(f"Erro ao encontrar colunas necessárias para Dim CR: {e}")
        return

    linhas_inseridas = 0
    for row in list(sheet.iter_rows(min_row=6, values_only=True)):
        cr_sap = str(row[idx_cr_sap]).strip() if row[idx_cr_sap] else None
        cod_cr = str(row[idx_cod_cr]).strip() if row[idx_cod_cr] else None
        if not cr_sap or cr_sap == 'None':
            continue
            
        cliente = str(row[idx_cliente]).strip() if row[idx_cliente] else None
        des_cr = str(row[idx_des_cr]).strip() if row[idx_des_cr] else None
        pais = str(row[idx_pais]).strip() if row[idx_pais] else None
        diretor = str(row[idx_diretor]).strip() if row[idx_diretor] else None
        gerente = str(row[idx_gerente]).strip() if row[idx_gerente] else None

        cursor.execute('''INSERT OR REPLACE INTO dim_cr
                         (Cod_Cr, CR_SAP, Cliente, Des_CR, Pais, Diretor, Gerente)
                         VALUES (?, ?, ?, ?, ?, ?, ?)''',
                       (cod_cr, cr_sap, cliente, des_cr, pais, diretor, gerente))
        linhas_inseridas += 1

    print(f"Dimensão CR carregada. Inseridas/Atualizadas: {linhas_inseridas}")


def run_etl():
    print(f"Iniciando ETL do arquivo {XLSX_PATH}")
    print(f"Usando banco de dados em: {DB_PATH}")
    conn = init_db()
    cursor = conn.cursor()
    ensure_forecast_oportunidades_schema(cursor)
    
    agora = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    try:
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
        
        # Carga DIM CR
        run_dim_cr_etl(wb, cursor, agora)
        
        sheet = wb["FORECAST"]
        
        header_row_idx, headers = find_forecast_header(sheet)
        print(f"FORECAST header encontrado em linha {header_row_idx}: {headers}")

        idx_map = {
            'chave_ek': find_header_index(headers, ['chave ek', 'chave_ek', 'ek', 'e k', 'chave']),
            'cr': find_header_index(headers, ['cr sap', 'cr_sap', 'cr', 'cod cr', 'cod_cr']),
            'data_criacao': find_header_index(headers, ['data criacao', 'data de criacao', 'data criacao', 'criacao']),
            'banco': find_header_index(headers, ['banco']),
            'pais': find_header_index(headers, ['pais', 'country']),
            'tipo_papel': find_header_index(headers, ['tipo papel', 'tipo de papel']),
            'owner': find_header_index(headers, ['owner', 'responsavel', 'proprietario']),
            'ger_comercial': find_header_index(headers, ['ger comercial', 'gerente comercial', 'gerencial', 'comercial']),
            'pratica': find_header_index(headers, ['pratica', 'pratica']),
            'produto': find_header_index(headers, ['produto']),
            'cliente': find_header_index(headers, ['cliente']),
            'subcliente': find_header_index(headers, ['subcliente']),
            'id_oportunidade': find_header_index(headers, ['id oportunidade', 'id_oportunidade', 'oportunidade', 'id']),
            'descricao_oportunidade': find_header_index(headers, ['descricao oportunidade', 'descricao', 'descricao de oportunidade', 'descricao oportunidade']),
            'status_comercial': find_header_index(headers, ['status comercial', 'status_comercial', 'status']),
            'moeda': find_header_index(headers, ['moeda']),
            'data_inicio_contrato': find_header_index(headers, ['data inicio contrato', 'data inicio', 'inicio contrato']),
            'data_fim_contrato': find_header_index(headers, ['data fim contrato', 'data fim', 'fim contrato']),
        }

        if idx_map['chave_ek'] is None:
            print('Aviso: chave_ek não encontrada por cabeçalho, usando fallback na coluna 140')
            idx_map['chave_ek'] = 140
        if idx_map['cr'] is None:
            print('Aviso: CR não encontrada por cabeçalho, usando fallback na coluna 24')
            idx_map['cr'] = 24

        lidas = 0
        carregadas = 0
        ignoradas = 0
        ignored_details = []
        cleaned_valores_keys = set()
        cr_4000020240 = {
            'found': False,
            'with_key': False,
            'row': None,
            'chave_ek': None,
        }

        for i, row in enumerate(sheet.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
            cr_raw = get_cell(row, idx_map['cr'])
            cr, cr2 = parse_cr_values(cr_raw)

            if cr == '4000020240' or cr2 == '4000020240':
                cr_4000020240['found'] = True
                cr_4000020240['chave_ek'] = get_cell(row, idx_map['chave_ek'])
                cr_4000020240['row'] = i
                if cr_4000020240['chave_ek']:
                    cr_4000020240['with_key'] = True

            chave_ek = get_cell(row, idx_map['chave_ek'])
            if not chave_ek:
                ignoradas += 1
                ignored_details.append(f"Linha {i} ignorada sem chave_ek (CR={cr})")
                continue

            lidas += 1

            try:
                opp = {
                    'chave_ek': str(chave_ek).strip(),
                    'data_criacao': parse_date(get_cell(row, idx_map['data_criacao'], row[0] if len(row) > 0 else None)),
                    'banco': str(get_cell(row, idx_map['banco'], row[1] if len(row) > 1 else None)).strip() if get_cell(row, idx_map['banco'], row[1] if len(row) > 1 else None) else None,
                    'pais': str(get_cell(row, idx_map['pais'], row[3] if len(row) > 3 else None)).strip() if get_cell(row, idx_map['pais'], row[3] if len(row) > 3 else None) else None,
                    'tipo_papel': str(get_cell(row, idx_map['tipo_papel'], row[4] if len(row) > 4 else None)).strip() if get_cell(row, idx_map['tipo_papel'], row[4] if len(row) > 4 else None) else None,
                    'owner': str(get_cell(row, idx_map['owner'], row[5] if len(row) > 5 else None)).strip() if get_cell(row, idx_map['owner'], row[5] if len(row) > 5 else None) else None,
                    'ger_comercial': str(get_cell(row, idx_map['ger_comercial'], row[6] if len(row) > 6 else None)).strip() if get_cell(row, idx_map['ger_comercial'], row[6] if len(row) > 6 else None) else None,
                    'pratica': str(get_cell(row, idx_map['pratica'], row[10] if len(row) > 10 else None)).strip() if get_cell(row, idx_map['pratica'], row[10] if len(row) > 10 else None) else None,
                    'produto': str(get_cell(row, idx_map['produto'], row[11] if len(row) > 11 else None)).strip() if get_cell(row, idx_map['produto'], row[11] if len(row) > 11 else None) else None,
                    'cliente': str(get_cell(row, idx_map['cliente'], row[13] if len(row) > 13 else None)).strip() if get_cell(row, idx_map['cliente'], row[13] if len(row) > 13 else None) else None,
                    'subcliente': str(get_cell(row, idx_map['subcliente'], row[14] if len(row) > 14 else None)).strip() if get_cell(row, idx_map['subcliente'], row[14] if len(row) > 14 else None) else None,
                    'id_oportunidade': str(get_cell(row, idx_map['id_oportunidade'], row[17] if len(row) > 17 else None)).strip() if get_cell(row, idx_map['id_oportunidade'], row[17] if len(row) > 17 else None) else None,
                    'descricao_oportunidade': str(get_cell(row, idx_map['descricao_oportunidade'], row[18] if len(row) > 18 else None)).strip() if get_cell(row, idx_map['descricao_oportunidade'], row[18] if len(row) > 18 else None) else None,
                    'status_comercial': str(get_cell(row, idx_map['status_comercial'], row[19] if len(row) > 19 else None)).strip() if get_cell(row, idx_map['status_comercial'], row[19] if len(row) > 19 else None) else None,
                    'cr': cr,
                    'cr2': cr2,
                    'moeda': str(get_cell(row, idx_map['moeda'], row[26] if len(row) > 26 else None)).strip() if get_cell(row, idx_map['moeda'], row[26] if len(row) > 26 else None) else None,
                    'data_inicio_contrato': parse_date(get_cell(row, idx_map['data_inicio_contrato'], row[29] if len(row) > 29 else None)),
                    'data_fim_contrato': parse_date(get_cell(row, idx_map['data_fim_contrato'], row[30] if len(row) > 30 else None)),
                }

                cursor.execute('''INSERT OR REPLACE INTO forecast_oportunidades 
                               (chave_ek, data_criacao, banco, pais, tipo_papel, owner, ger_comercial,
                                pratica, produto, cliente, subcliente, id_oportunidade, descricao_oportunidade,
                                status_comercial, cr, cr2, moeda, data_inicio_contrato, data_fim_contrato, semana_carga, arquivo_origem)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                               (opp['chave_ek'], opp['data_criacao'], opp['banco'], opp['pais'],
                                opp['tipo_papel'], opp['owner'], opp['ger_comercial'], opp['pratica'],
                                opp['produto'], opp['cliente'], opp['subcliente'], opp['id_oportunidade'],
                                opp['descricao_oportunidade'], opp['status_comercial'], opp['cr'], opp['cr2'], opp['moeda'],
                                opp['data_inicio_contrato'], opp['data_fim_contrato'], agora, XLSX_PATH))

                # Limpa valores da chave apenas uma vez por execução do ETL.
                # Evita apagar valores válidos quando a mesma chave_ek aparece em linhas duplicadas.
                if opp['chave_ek'] not in cleaned_valores_keys:
                    cursor.execute('DELETE FROM forecast_valores WHERE chave_ek = ?', (opp['chave_ek'],))
                    cleaned_valores_keys.add(opp['chave_ek'])

                meses = ['01','02','03','04','05','06','07','08','09','10','11','12']
                for m_idx, mes in enumerate(meses):
                    v_rl_25 = parse_float(get_cell(row, 34 + m_idx)) if len(row) > 34 + m_idx else 0.0
                    v_rb_25 = parse_float(get_cell(row, 60 + m_idx)) if len(row) > 60 + m_idx else 0.0
                    if v_rl_25 != 0 or v_rb_25 != 0:
                        mes_ref = f'2025-{mes}'
                        cursor.execute(
                            'UPDATE forecast_valores SET valor_rl = COALESCE(valor_rl, 0) + ?, valor_rb = COALESCE(valor_rb, 0) + ?, semana_carga = ? '
                            'WHERE chave_ek = ? AND cenario = ? AND mes_ref = ?',
                            (v_rl_25, v_rb_25, agora, opp['chave_ek'], 'Forecast', mes_ref)
                        )
                        if cursor.rowcount == 0:
                            cursor.execute(
                                'INSERT INTO forecast_valores (chave_ek, cenario, mes_ref, valor_rl, valor_rb, semana_carga) VALUES (?,?,?,?,?,?)',
                                (opp['chave_ek'], 'Forecast', mes_ref, v_rl_25, v_rb_25, agora)
                            )

                for m_idx, mes in enumerate(meses):
                    v_rl_26 = parse_float(get_cell(row, 47 + m_idx)) if len(row) > 47 + m_idx else 0.0
                    v_rb_26 = parse_float(get_cell(row, 125 + m_idx)) if len(row) > 125 + m_idx else 0.0
                    if v_rl_26 != 0 or v_rb_26 != 0:
                        mes_ref = f'2026-{mes}'
                        cursor.execute(
                            'UPDATE forecast_valores SET valor_rl = COALESCE(valor_rl, 0) + ?, valor_rb = COALESCE(valor_rb, 0) + ?, semana_carga = ? '
                            'WHERE chave_ek = ? AND cenario = ? AND mes_ref = ?',
                            (v_rl_26, v_rb_26, agora, opp['chave_ek'], 'Forecast', mes_ref)
                        )
                        if cursor.rowcount == 0:
                            cursor.execute(
                                'INSERT INTO forecast_valores (chave_ek, cenario, mes_ref, valor_rl, valor_rb, semana_carga) VALUES (?,?,?,?,?,?)',
                                (opp['chave_ek'], 'Forecast', mes_ref, v_rl_26, v_rb_26, agora)
                            )

                carregadas += 1
            except Exception as e:
                ignored_details.append(f"Linha {i} parse falhou: {e}")
                ignoradas += 1
                continue

        if ignored_details:
            print('Linhas ignoradas detalhadas:')
            for msg in ignored_details[:30]:
                print(' -', msg)
            if len(ignored_details) > 30:
                print(f' - ... mais {len(ignored_details) - 30} linhas ignoradas')

        if cr_4000020240['found']:
            if cr_4000020240['with_key']:
                print(f"CR 4000020240 encontrado na linha {cr_4000020240['row']} e possui chave_ek {cr_4000020240['chave_ek']}")
            else:
                print(f"CR 4000020240 encontrado na linha {cr_4000020240['row']}, mas sem chave_ek")
        else:
            print('CR 4000020240 não encontrado no source FORECAST.')

        message = f"Carga finalizada com sucesso; cr_4000020240_found={cr_4000020240['found']}; cr_4000020240_with_key={cr_4000020240['with_key']}"
        cursor.execute("INSERT INTO etl_log (arquivo, aba, linhas_lidas, linhas_carregadas, linhas_ignoradas, status, mensagem, executado_em) VALUES (?,?,?,?,?,?,?,?)",
                       (XLSX_PATH, "FORECAST", lidas, carregadas, ignoradas, "SUCESSO", message, agora))
        conn.commit()
        print(f"ETL Concluído. Lidas: {lidas}, Carregadas: {carregadas}, Ignoradas: {ignoradas}")

    except Exception as e:
        cursor.execute("INSERT INTO etl_log (arquivo, aba, linhas_lidas, linhas_carregadas, linhas_ignoradas, status, mensagem, executado_em) VALUES (?,?,?,?,?,?,?,?)",
                       (XLSX_PATH, "FORECAST", 0, 0, 0, "ERRO", str(e), agora))
        conn.commit()
        print(f"Erro fatal no ETL: {e}")
    finally:
        conn.close()

if __name__ == "__main__":
    run_etl()
