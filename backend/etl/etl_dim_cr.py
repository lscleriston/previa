import os
import sqlite3
import openpyxl

DB_PATH = os.environ.get(
    "DB_PATH",
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "db", "previadb.db")
)
XLSX_PATH = os.environ.get(
    "XLSX_PATH",
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "raw", "Forecast Semanal 2026 - Abril.xlsx")
)

def run_dim_cr_etl():
    print("Iniciando ETL da tabela Dimensão CR...")
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(dim_cr)")
    existing_cols = [row[1] for row in cursor.fetchall()]
    if 'CR_SAP' not in existing_cols:
        cursor.execute("ALTER TABLE dim_cr ADD COLUMN CR_SAP TEXT")
        cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_dim_cr_cr_sap ON dim_cr(CR_SAP)")
    
    try:
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
        # Identificar aba que começa com "Prévia" e termina com "MSP" ou contém "MSP"
        sheet_name = None
        for s in wb.sheetnames:
            if s.startswith('Prévia') and 'MSP' in s:
                sheet_name = s
                break
        
        if not sheet_name:
            print("Aba de Prévia não encontrada.")
            return

        print(f"Lendo dados da aba: {sheet_name}")
        sheet = wb[sheet_name]
        
        # Identificar os índices baseado na linha 5
        header_row = list(sheet.iter_rows(min_row=5, max_row=5, values_only=True))[0]
        
        try:
            idx_cr_sap = header_row.index('CR SAP')
            idx_cod_cr = header_row.index('Cod_Cr')
            idx_cliente = header_row.index('Cliente')
            idx_des_cr = header_row.index('Des_CR')
            idx_pais = header_row.index('Pais')
            idx_diretor = header_row.index('Diretor')
            idx_gerente = header_row.index('Gerente')
        except ValueError as e:
            print(f"Erro ao encontrar colunas necessárias: {e}")
            return

        linhas_lidas = 0
        linhas_inseridas = 0
        for row in sheet.iter_rows(min_row=6, values_only=True):
            linhas_lidas += 1
            cr_sap = str(row[idx_cr_sap]).strip() if row[idx_cr_sap] else None
            cod_cr = str(row[idx_cod_cr]).strip() if row[idx_cod_cr] else None
            if not cr_sap or cr_sap == 'None':
                continue
                
            cliente = str(row[idx_cliente]).strip() if row[idx_cliente] else None
            des_cr = str(row[idx_des_cr]).strip() if row[idx_des_cr] else None
            pais = str(row[idx_pais]).strip() if row[idx_pais] else None
            diretor = str(row[idx_diretor]).strip() if row[idx_diretor] else None
            gerente = str(row[idx_gerente]).strip() if row[idx_gerente] else None

            # Insert/Update (Upsert)
            cursor.execute('''INSERT OR REPLACE INTO dim_cr
                             (Cod_Cr, CR_SAP, Cliente, Des_CR, Pais, Diretor, Gerente)
                             VALUES (?, ?, ?, ?, ?, ?, ?)''',
                           (cod_cr, cr_sap, cliente, des_cr, pais, diretor, gerente))
            linhas_inseridas += 1

        conn.commit()
        print(f"ETL Concluído. Lidas: {linhas_lidas}, Inseridas/Atualizadas: {linhas_inseridas}")

    except Exception as e:
        print(f"Erro durante ETL da dimensão CR: {e}")
    finally:
        conn.close()

if __name__ == "__main__":
    run_dim_cr_etl()