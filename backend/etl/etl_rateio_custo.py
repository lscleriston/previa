import sqlite3
import os
import openpyxl
from datetime import datetime

DB_PATH = os.environ.get(
    "DB_PATH",
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "db", "previadb.db")
)
XLSX_PATH = os.environ.get(
    "XLSX_PATH",
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "raw", "Forecast Semanal 2026 - Abril.xlsx")
)

SHEET_NAME = "RATEIO CUSTO DIRETO MSP AUTOMAT"


def parse_float(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except Exception:
        try:
            text = str(val).strip().replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
            return float(text)
        except Exception:
            return 0.0


def normalize_text(val):
    if val is None:
        return None
    text = str(val).strip()
    return text if text else None


def normalize_cr(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if float(val).is_integer():
            return str(int(val))
        return str(val).strip()
    text = str(val).strip()
    if text.endswith('.0') and text[:-2].isdigit():
        return text[:-2]
    return text


def parse_mes(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m')
    text = str(val).strip()
    if not text:
        return None
    if len(text) >= 7 and text[4] == '-':
        return text[:7]
    if '/' in text:
        parts = [p for p in text.replace(' ', '').split('/') if p]
        if len(parts) >= 2:
            try:
                month = int(parts[-2])
                year = int(parts[-1])
                return f"{year:04d}-{month:02d}"
            except Exception:
                pass
    return text


def run_etl():
    print(f"Iniciando ETL de Rateio Direto a partir da aba {SHEET_NAME}...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"Aba {SHEET_NAME} não encontrada.")
        return

    sheet = wb[SHEET_NAME]
    header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)
    if not header_row:
        print("Cabeçalho não encontrado na aba de rateio.")
        return

    header_map = {}
    for idx, heading in enumerate(header_row):
        if not heading:
            continue
        header_map[str(heading).strip().upper()] = idx

    required = ["COD CR", "RATEIO", "CR DE ENVIO", "DESCR CR DE ENVIO", "RESPONSÁVEL CR ENVIO", "MÊS"]
    missing = [col for col in required if col not in header_map]
    if missing:
        print(f"Cabeçalhos ausentes na aba de rateio: {missing}")
        return

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS rateio_automatico (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cr TEXT,
            mes_ref TEXT,
            rateio REAL,
            cr_envio TEXT,
            desc_cr_envio TEXT,
            gestor_cr_envio TEXT,
            cliente TEXT,
            rl_rateio TEXT,
            descricao TEXT,
            aba_origem TEXT
        )
    ''')
    cursor.execute("DELETE FROM rateio_automatico")
    count = 0

    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        cod_cr = normalize_cr(row[header_map["COD CR"]])
        if not cod_cr:
            continue

        rateio = parse_float(row[header_map["RATEIO"]])
        if rateio == 0:
            continue
        rateio = -abs(rateio)

        cr_envio = normalize_cr(row[header_map["CR DE ENVIO"]])
        desc_cr_envio = normalize_text(row[header_map["DESCR CR DE ENVIO"]])
        gestor_cr_envio = normalize_text(row[header_map["RESPONSÁVEL CR ENVIO"]])
        mes_ref = parse_mes(row[header_map["MÊS"]])
        cliente = normalize_text(row[header_map.get("CLIENTE")]) if "CLIENTE" in header_map else None
        rl_rateio = normalize_text(row[header_map.get("RL RATEIO")]) if "RL RATEIO" in header_map else None

        descricao = cliente or rl_rateio or "Rateio"

        cursor.execute('''
            INSERT INTO rateio_automatico (
                cr, mes_ref, rateio, cr_envio, desc_cr_envio,
                gestor_cr_envio, cliente, rl_rateio, descricao, aba_origem
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            cod_cr,
            mes_ref,
            rateio,
            cr_envio,
            desc_cr_envio,
            gestor_cr_envio,
            cliente,
            rl_rateio,
            descricao,
            SHEET_NAME
        ))
        count += 1

    conn.commit()
    conn.close()
    print(f"ETL Rateio Direto concluído. {count} registros inseridos.")


if __name__ == "__main__":
    run_etl()
