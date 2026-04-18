import sqlite3
import os
import openpyxl
import re

DB_PATH = os.environ.get(
    "DB_PATH",
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "db", "previadb.db")
)
XLSX_PATH = os.environ.get(
    "XLSX_PATH",
    os.path.join(os.path.dirname(__file__), "..", "..", "data", "raw", "Forecast Semanal 2026 - Abril.xlsx")
)

def create_table_if_not_exists(conn):
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS orcamento_previa (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mes_ref TEXT,
            cr TEXT,
            cliente TEXT,
            des_cr TEXT,
            pais TEXT,
            diretor TEXT,
            gerente TEXT,
            owner TEXT,
            categoria_despesa TEXT,
            valor_plano REAL
        )
    ''')
    conn.commit()

def parse_float(val):
    if val is None:
        return 0.0
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).replace('R$', '').replace(' ', '').replace(',', '.')
        return float(s)
    except:
        return 0.0

def run_etl():
    print(f"Iniciando ETL Orçado x Prévia a partir de {XLSX_PATH}")
    
    if not os.path.exists(XLSX_PATH):
        print(f"Arquivo não encontrado: {XLSX_PATH}")
        return
        
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    
    # Encontrar a aba "Prévia {Mês} {Ano} - MSP"
    sheet_name = None
    for s in wb.sheetnames:
        if s.startswith('Prévia') and 'MSP' in s:
            sheet_name = s
            break
            
    if not sheet_name:
        print("Aba de Prévia não encontrada.")
        return
        
    print(f"Aba encontrada: {sheet_name}")
    
    # Extrair mês e ano do nome da aba (ex: "Prévia Abr 2026 - MSP")
    # Para simplificar e bater com os dados, vamos assumir o mes como 2026-04 pelo nome do arquivo ou da aba
    mes_ref = '2026-04' # Fixando para este exemplo conforme outros ETLs
    
    sheet = wb[sheet_name]
    
    # Buscar os cabeçalhos na linha 4 (Categorias)
    row4 = [col.value for col in sheet[4]]
    # Linha 5 tem os nomes das colunas
    row5 = [col.value for col in sheet[5]]
    
    # Mapear os indices das colunas onde row5 tem nome igual "Plano 26" ou onde tem "Plano 26"
    # A categoria estará associada ao índice em row4. A categoria em row4 geralmente está agrupada, e pode aparecer antes do "Plano 26".
    # Então vamos iterar e salvar a categoria atual.
    
    categorias_plano = [] # Lista de tuplas (index, categoria_name)
    current_category = "Geral"
    
    for idx, (cat_val, col_name) in enumerate(zip(row4, row5)):
        if cat_val is not None:
            current_category = str(cat_val).strip()
            
        if col_name is not None and ('Plano' in str(col_name) or 'plano' in str(col_name).lower()):
            categorias_plano.append((idx, current_category))

    print(f"Encontradas {len(categorias_plano)} colunas de Plano 26: {categorias_plano}")
    
    conn = sqlite3.connect(DB_PATH)
    create_table_if_not_exists(conn)
    
    conn.execute("DELETE FROM orcamento_previa WHERE mes_ref = ?", (mes_ref,))
    conn.commit()
    
    cursor = conn.cursor()
    count = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=6, values_only=True)):
        cr = str(row[2]) if row[2] is not None else None
        if not cr or cr == 'None' or cr == 'Total Result':
            continue
            
        cliente = str(row[3]) if row[3] is not None else None
        des_cr = str(row[4]) if row[4] is not None else None
        pais = str(row[5]) if row[5] is not None else None
        diretor = str(row[6]) if row[6] is not None else None
        gerente = str(row[7]) if row[7] is not None else None
        owner = str(row[8]) if row[8] is not None else None
        
        for idx, categoria in categorias_plano:
            val = parse_float(row[idx])
            if val != 0.0:
                cursor.execute('''
                    INSERT INTO orcamento_previa (
                        mes_ref, cr, cliente, des_cr, pais, diretor, gerente, owner, categoria_despesa, valor_plano
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    mes_ref, cr, cliente, des_cr, pais, diretor, gerente, owner, categoria, val
                ))
                count += 1
                
    conn.commit()
    conn.close()
    print(f"ETL Orçado concluído. {count} registros de orçamentos (Planos) inseridos.")

if __name__ == '__main__':
    run_etl()