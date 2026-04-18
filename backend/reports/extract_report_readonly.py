import os
import openpyxl
import json

xls_path = os.path.join(os.path.dirname(__file__), "..", "..", "data", "raw", "Forecast Semanal 2026 - Abril.xlsx")
print("Loading workbook (read-only)...")
wb = openpyxl.load_workbook(xls_path, data_only=True, read_only=True)
sheet = wb["FORECAST"]

head_row = None
for i, row in enumerate(sheet.iter_rows(max_row=70, values_only=True), start=1):
    if i == 65:
        print(f"Row 65: {row[:5]}")
    # Finding header row dynamically or checking row 65
    # Since we need to answer "Confirme isso (linha 65)", we MUST check row 65.
    
    if i == 64:
        row64 = row
    if i == 65:
        row65 = row

# We will just assume row 65 as instructed by user prompt context, wait, 
# The prompt says: "A linha 65 deve conter os nomes das colunas. Confirme isso."
# Let's extract row 65 anyway to see its contents fully.

headers = [str(c).strip() if c is not None else "" for c in row65]
print(f"Colunas encontradas na linha 65: {len(headers)}")

# Salva apenas cabeçalhos não-vazios e o seu índice (1-based)
cols = []
for i, h in enumerate(headers):
    # Pode haver alguns vazios
    cols.append({'idx': i+1, 'col_letter': openpyxl.utils.get_column_letter(i+1), 'name': h[:50]})  # keep it short if it's data

sample_rows = []
for i, row in enumerate(sheet.iter_rows(min_row=66, max_row=70, values_only=True), start=66):
    row_data = [str(c).strip() if c is not None else "" for c in row]
    sample_rows.append(row_data)

result = {
    'row65_headers': cols,
    'samples': sample_rows
}

output_path = os.path.join(os.path.dirname(__file__), "..", "..", "data", "reports", "report_etapa1_readonly.json")
with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"Relatório salvo em {output_path}")
