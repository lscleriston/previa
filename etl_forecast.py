import sqlite3
import openpyxl
import datetime
import os
import glob

DB_PATH = os.environ.get("DB_PATH", "previadb.db")
XLSX_PATH = os.environ.get("XLSX_PATH", "Forecast Semanal 2026 - Abril.xlsx")

def clear_db():
    pass

def init_db():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.executescript("""
    CREATE TABLE IF NOT EXISTS forecast_oportunidades (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chave_ek TEXT UNIQUE,
        data_criacao DATE,
        banco TEXT,
        ano INTEGER,
        pais TEXT,
        tipo_papel TEXT,
        gerente TEXT,
        owner TEXT,
        ger_comercial TEXT,
        operacao TEXT,
        ger_operacao TEXT,
        pre_vendas TEXT,
        pratica TEXT,
        produto TEXT,
        id_empresa TEXT,
        cliente TEXT,
        cliente_ge TEXT,
        subcliente TEXT,
        novo_cliente BOOLEAN,
        industria TEXT,
        id_oportunidade TEXT,
        descricao_oportunidade TEXT,
        status_comercial TEXT,
        status_comercial_det TEXT,
        metodologia TEXT,
        tipo_opp TEXT,
        cr TEXT,
        semana_fechamento TEXT,
        cr_oi TEXT,
        ordem_interna TEXT,
        moeda TEXT,
        consideracao TEXT,
        vigencia TEXT,
        contrato TEXT,
        data_inicio_contrato DATE,
        data_fim_contrato DATE,
        mes_reajuste TEXT,
        risco TEXT,
        deducao REAL,
        pct_ponderacao REAL,
        semana_carga DATETIME,
        arquivo_origem TEXT
    );

    CREATE TABLE IF NOT EXISTS forecast_valores (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chave_ek TEXT,
        cenario TEXT,
        mes_ref TEXT,
        valor_rl REAL,
        valor_rb REAL,
        semana_carga DATETIME,
        FOREIGN KEY (chave_ek) REFERENCES forecast_oportunidades(chave_ek)
    );

    CREATE TABLE IF NOT EXISTS etl_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        arquivo TEXT,
        aba TEXT,
        linhas_lidas INTEGER,
        linhas_carregadas INTEGER,
        linhas_ignoradas INTEGER,
        status TEXT,
        mensagem TEXT,
        executado_em DATETIME
    );
    """)
    conn.commit()
    return conn

def parse_date(val):
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    return None

def parse_float(val):
    if not val: return 0.0
    try:
        return float(val)
    except:
        return 0.0

def run_dim_cr_etl(wb, cursor, agora):
    print("Iniciando carga da Dimensão CR...")
    sheet_name = None
    for s in wb.sheetnames:
        if s.startswith('Prévia') and 'MSP' in s:
            sheet_name = s
            break
            
    if not sheet_name:
        print("Aba de Prévia não encontrada para carga da dim_cr.")
        return

    print(f"Lendo Dim CR da aba: {sheet_name}")
    sheet = wb[sheet_name]
    cursor.execute("PRAGMA table_info(dim_cr)")
    existing_cols = [row[1] for row in cursor.fetchall()]
    if 'CR_SAP' not in existing_cols:
        cursor.execute("ALTER TABLE dim_cr ADD COLUMN CR_SAP TEXT")
        cursor.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_dim_cr_cr_sap ON dim_cr(CR_SAP)")
    
    header_row = list(sheet.iter_rows(min_row=5, max_row=5, values_only=True))[0]
    
    try:
        idx_cr_sap = header_row.index('CR SAP')
        idx_cod_cr = header_row.index('Cod_Cr')
        idx_cliente = header_row.index('Cliente')
        idx_des_cr = header_row.index('Des_CR')
        idx_pais = header_row.index('Pais')
        idx_diretor = header_row.index('Diretor')
        idx_gerente = header_row.index('Gerente')
    except ValueError as e:
        print(f"Erro ao encontrar colunas necessárias para Dim CR: {e}")
        return

    linhas_inseridas = 0
    for row in list(sheet.iter_rows(min_row=6, values_only=True)):
        cr_sap = str(row[idx_cr_sap]).strip() if row[idx_cr_sap] else None
        cod_cr = str(row[idx_cod_cr]).strip() if row[idx_cod_cr] else None
        if not cr_sap or cr_sap == 'None':
            continue
            
        cliente = str(row[idx_cliente]).strip() if row[idx_cliente] else None
        des_cr = str(row[idx_des_cr]).strip() if row[idx_des_cr] else None
        pais = str(row[idx_pais]).strip() if row[idx_pais] else None
        diretor = str(row[idx_diretor]).strip() if row[idx_diretor] else None
        gerente = str(row[idx_gerente]).strip() if row[idx_gerente] else None

        cursor.execute('''INSERT OR REPLACE INTO dim_cr
                         (Cod_Cr, CR_SAP, Cliente, Des_CR, Pais, Diretor, Gerente)
                         VALUES (?, ?, ?, ?, ?, ?, ?)''',
                       (cod_cr, cr_sap, cliente, des_cr, pais, diretor, gerente))
        linhas_inseridas += 1

    print(f"Dimensão CR carregada. Inseridas/Atualizadas: {linhas_inseridas}")


def run_etl():
    print(f"Iniciando ETL do arquivo {XLSX_PATH}")
    conn = init_db()
    cursor = conn.cursor()
    
    agora = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    try:
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
        
        # Carga DIM CR
        run_dim_cr_etl(wb, cursor, agora)
        
        sheet = wb["FORECAST"]
        
        # Como descoberto na Etapa 1, os dados reais parecem começar nas primeiras linhas ou na 66 dependendo das ocultas.
        # Estamos iterando a partir da linha 4
        
        lidas = 0
        carregadas = 0
        ignoradas = 0
        
        for i, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
            # A chave única está na coluna 141 (EK) ou formamos ela
            chave_ek = row[140] if len(row) > 140 and row[140] else None
            if not chave_ek:
                ignoradas += 1
                continue
                
            lidas += 1
            
            # Mapeamento Básico Identificado (Ajustar com testes reais se colunas variarem)
            try:
                opp = {
                    "chave_ek": str(chave_ek).strip(),
                    "data_criacao": parse_date(row[0]),
                    "banco": str(row[1]).strip() if row[1] else None,
                    "pais": str(row[3]).strip() if row[3] else None,
                    "tipo_papel": str(row[4]).strip() if row[4] else None,
                    "owner": str(row[5]).strip() if row[5] else None,
                    "ger_comercial": str(row[6]).strip() if row[6] else None,
                    "pratica": str(row[10]).strip() if row[10] else None,
                    "produto": str(row[11]).strip() if row[11] else None,
                    "cliente": str(row[13]).strip() if row[13] else None,
                    "subcliente": str(row[14]).strip() if row[14] else None,
                    "id_oportunidade": str(row[17]).strip() if row[17] else None,
                    "descricao_oportunidade": str(row[18]).strip() if row[18] else None,
                    "status_comercial": str(row[19]).strip() if row[19] else None,
                    "cr": str(row[24]).split('-')[0].strip() if len(row)>24 and row[24] else None,
                    "moeda": str(row[26]).strip() if len(row)>26 and row[26] else None,
                    "data_inicio_contrato": parse_date(row[29]) if len(row)>29 else None,
                    "data_fim_contrato": parse_date(row[30]) if len(row)>30 else None,
                }
                
                # Inserir Oportunidade
                cursor.execute('''INSERT OR REPLACE INTO forecast_oportunidades 
                               (chave_ek, data_criacao, banco, pais, tipo_papel, owner, ger_comercial,
                                pratica, produto, cliente, subcliente, id_oportunidade, descricao_oportunidade,
                                status_comercial, cr, moeda, data_inicio_contrato, data_fim_contrato, semana_carga, arquivo_origem)
                               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                               (opp['chave_ek'], opp['data_criacao'], opp['banco'], opp['pais'],
                                opp['tipo_papel'], opp['owner'], opp['ger_comercial'], opp['pratica'],
                                opp['produto'], opp['cliente'], opp['subcliente'], opp['id_oportunidade'],
                                opp['descricao_oportunidade'], opp['status_comercial'], opp['cr'], opp['moeda'],
                                opp['data_inicio_contrato'], opp['data_fim_contrato'], agora, XLSX_PATH))
                
                # Deletar valores antigos
                cursor.execute('DELETE FROM forecast_valores WHERE chave_ek = ?', (opp['chave_ek'],))
                
                # 2025: RL (34 a 45) e RB (60 a 71)
                meses = ['01','02','03','04','05','06','07','08','09','10','11','12']
                for m_idx, mes in enumerate(meses):
                    v_rl_25 = parse_float(row[34 + m_idx]) if len(row) > 45 else 0.0
                    v_rb_25 = parse_float(row[60 + m_idx]) if len(row) > 71 else 0.0
                    
                    if v_rl_25 != 0 or v_rb_25 != 0:
                        cursor.execute('INSERT INTO forecast_valores (chave_ek, cenario, mes_ref, valor_rl, valor_rb, semana_carga) VALUES (?,?,?,?,?,?)',
                                       (opp['chave_ek'], 'Forecast', f'2025-{mes}', v_rl_25, v_rb_25, agora))

                # 2026: RL nas colunas AV:BG (índice 47 a 58) e RB nas colunas DV:EG (índice 125 a 136)
                for m_idx, mes in enumerate(meses):
                    v_rl_26 = parse_float(row[47 + m_idx]) if len(row) > 58 else 0.0
                    v_rb_26 = parse_float(row[125 + m_idx]) if len(row) > 136 else 0.0
                    
                    if v_rl_26 != 0 or v_rb_26 != 0:
                        cursor.execute('INSERT INTO forecast_valores (chave_ek, cenario, mes_ref, valor_rl, valor_rb, semana_carga) VALUES (?,?,?,?,?,?)',
                                       (opp['chave_ek'], 'Forecast', f'2026-{mes}', v_rl_26, v_rb_26, agora))
                
                carregadas += 1
            except Exception as e:
                print(f"Erro ao parsear linha {i}: {e}")
                ignoradas += 1

        cursor.execute("INSERT INTO etl_log (arquivo, aba, linhas_lidas, linhas_carregadas, linhas_ignoradas, status, mensagem, executado_em) VALUES (?,?,?,?,?,?,?,?)",
                       (XLSX_PATH, "FORECAST", lidas, carregadas, ignoradas, "SUCESSO", "Carga finalizada com sucesso", agora))
        conn.commit()
        print(f"ETL Concluído. Lidas: {lidas}, Carregadas: {carregadas}, Ignoradas: {ignoradas}")

    except Exception as e:
        cursor.execute("INSERT INTO etl_log (arquivo, aba, linhas_lidas, linhas_carregadas, linhas_ignoradas, status, mensagem, executado_em) VALUES (?,?,?,?,?,?,?,?)",
                       (XLSX_PATH, "FORECAST", 0, 0, 0, "ERRO", str(e), agora))
        conn.commit()
        print(f"Erro fatal no ETL: {e}")
    finally:
        conn.close()

if __name__ == "__main__":
    run_etl()
